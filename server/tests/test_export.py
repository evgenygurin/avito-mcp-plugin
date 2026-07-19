"""Тесты модуля экспорта (exporter)."""

import json
import tempfile
from pathlib import Path

import pytest

from io import BytesIO

from avito_mcp_server.export.exporter import (
    export_listings,
    to_csv_str,
    to_json_str,
    to_xlsx_bytes,
)
from avito_mcp_server.models import Listing


@pytest.fixture()
def listings() -> list[Listing]:
    return [
        Listing(
            id=1,
            title="квартира",
            price=5000000,
            address="NN",
            url="https://www.avito.ru/x_1",
            published_at=1700000000,
            views=42,
        ),
        Listing(id=2, title="гараж", price=900000, address="NN"),
    ]


class TestXlsx:
    def test_returns_bytes(self, listings: list[Listing]) -> None:
        data = to_xlsx_bytes(listings)
        assert isinstance(data, bytes)
        assert len(data) > 0

    def test_export_to_file(self, listings: list[Listing]) -> None:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            path = f.name
        try:
            content, written = export_listings(listings, "xlsx", path)
            assert content == ""
            assert written == path
            assert Path(path).exists()
        finally:
            Path(path).unlink(missing_ok=True)

    def test_export_to_base64(self, listings: list[Listing]) -> None:
        content, written = export_listings(listings, "xlsx")
        assert written is None
        import base64

        decoded = base64.b64decode(content)
        assert len(decoded) > 0


class TestJson:
    def test_returns_json_str(self, listings: list[Listing]) -> None:
        data = to_json_str(listings)
        parsed = json.loads(data)
        assert len(parsed) == 2
        assert parsed[0]["id"] == 1

    def test_export_to_file(self, listings: list[Listing]) -> None:
        with tempfile.NamedTemporaryFile(suffix=".json", delete=False) as f:
            path = f.name
        try:
            content, written = export_listings(listings, "json", path)
            assert written == path
            assert json.loads(Path(path).read_text())[0]["id"] == 1
        finally:
            Path(path).unlink(missing_ok=True)


class TestCsv:
    def test_returns_csv_str(self, listings: list[Listing]) -> None:
        data = to_csv_str(listings)
        assert "id,title,price" in data
        assert "квартира" in data

    def test_export_to_file(self, listings: list[Listing]) -> None:
        with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as f:
            path = f.name
        try:
            content, written = export_listings(listings, "csv", path)
            assert written == path
        finally:
            Path(path).unlink(missing_ok=True)


def test_unknown_format_raises(listings: list[Listing]) -> None:
    with pytest.raises(ValueError, match="неподдерживаемый формат"):
        export_listings(listings, "pdf")


def test_xlsx_survives_control_characters() -> None:
    # В описаниях Avito встречаются управляющие символы; openpyxl на них падает
    # с IllegalCharacterError, и весь экспорт уходит в ошибку.
    from openpyxl import load_workbook

    listings = [Listing(id=1, title="квартира\x07", description="текст\x0bещё")]
    data = to_xlsx_bytes(listings)

    ws = load_workbook(BytesIO(data)).active
    assert "\x07" not in ws["B2"].value
    assert ws["B2"].value == "квартира"


def test_xlsx_neutralizes_formulas() -> None:
    # Заголовок объявления подконтролен продавцу: '=' в начале превращает ячейку
    # в формулу, которая выполнится при открытии файла.
    from openpyxl import load_workbook

    listings = [Listing(id=1, title='=HYPERLINK("http://evil","click")')]
    ws = load_workbook(BytesIO(to_xlsx_bytes(listings))).active

    assert ws["B2"].data_type == "s", "значение не должно быть формулой"
    assert not str(ws["B2"].value).startswith("=")


def test_csv_neutralizes_formulas() -> None:
    rows = to_csv_str([Listing(id=1, title="+1234")]).splitlines()
    assert not rows[1].split(",")[1].startswith("+")


def test_xlsx_writes_numbers_as_numbers() -> None:
    # Цена и id строками ломают сортировку и формулы в самом Excel.
    from openpyxl import load_workbook

    ws = load_workbook(BytesIO(to_xlsx_bytes([Listing(id=7, title="x", price=50000.0)]))).active
    assert ws["A2"].value == 7
    assert ws["C2"].value == 50000


def test_xlsx_without_path_rejects_huge_payload() -> None:
    # Base64 бинарника уходит прямо в ответ MCP-тулзы и съедает контекст модели.
    # Небольшой файл пропускаем, крупный — просим указать path.
    import uuid

    # Случайный текст, чтобы zip не сжал таблицу до нескольких килобайт.
    many = [
        Listing(id=i, title=uuid.uuid4().hex * 8, description=uuid.uuid4().hex * 8)
        for i in range(3000)
    ]
    with pytest.raises(ValueError, match="path"):
        export_listings(many, "xlsx")


def test_small_xlsx_without_path_still_works() -> None:
    content, written_path = export_listings([Listing(id=1, title="x")], "xlsx")
    assert written_path is None
    assert content
